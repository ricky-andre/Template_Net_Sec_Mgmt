
import re
import time
import datetime
import os.path
from ipaddress import IPv4Address
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill

# Where the output of the configurations check is written, here for every profile
# you can find the list of devices, for each of them you can find which config commands
# are present or not
xls_cfg_miss = './Cfg_Check.xlsx'
xls_ip_devices = './IpDevices_list.xlsx'
cfg_root_dir = 'root_path_to_config_files'
# File that contains the configurations to be checked for every profile, and what to do in case
# of mismatches (add a command in case something is missing, or remove/change a command)
cfg_check_cmd = './cfg_cmd_check.xlsx'

# list of the files to be taken as 'template' examples
cfg_template = {    
    "PROFILE_1": cfg_root_dir + "Backbone/pe_1.txt",
    "PROFILE_2": cfg_root_dir + "Backbone/rr_1.txt",
    "PROFILE_3": cfg_root_dir + "Backbone/p.txt"
}

# generate the proposed config template by reading the configs, should be manually revised
gen_template = False
# generate the output matrix with the template commands, and "what is present where"
gen_excel = True
# check what needs to be fixed, using columns 'B' and 'C' in cfg_check_cmd excel file
fix_cfg = False


profiles_filter = '.*'
dev_filter = '.*'
# command lines that are 'sons' of other lines, are separated in this way
line_break = '@@@'

commands = {}

def get_line (line):
    # This function returns the config line, changing it if necessary to reflect the specific requirements.
    # In case the line needs to be skipped, 'None' is returned.
    if re.search("^\s+!", line):
        return None
    elif re.search("^(vrf definition|mpls ldp neighbor.*password|hostname)", line):
        return None
    elif re.search("^!", line):
        return None
    elif re.search("^(show run|Building config|Current config|version|boot)", line):
        return None
    elif re.search("set uuid", line):
        return None

    # Lines that need to be changed to be properly matched ...
    if re.search("enable secret \d+", line):
        line = re.search("(enable secret \d+)", line).group(1)
    elif re.search("server-private.*key 7 ", line):
        line = re.search('(.*key 7)', line).group(1)
    elif re.search("password 7", line):
        line = re.search('(.*password 7)', line).group(1)
    elif re.search("key-string \d+", line):
        line = re.search('(.*key-string \d+)', line).group(1)
    elif re.search("license udi pid.*sn", line):
        line = "license udi pid.*sn"
    elif re.search("^username.*secret \d+", line):
        line = re.search("(username.*secret \d+)", line).group(1)
    elif re.search("^ntp authentication-key 10 md5", line):
        line = "ntp authentication-key 10 md5"
    elif re.search("crypto pki trustpoint TP-self-signed-", line):
        line = "crypto pki trustpoint TP-self-signed-"
    elif re.search("Self-Signed-Certificate-", line):
        line = re.search("(.*Self-Signed-Certificate-)", line).group(1)
    elif re.search("TP-self-signed-", line):
        line = re.search("(.*TP-self-signed-)", line).group(1)
    return line

def update_template_cmd (filename, target = False):
    """
    We now need to parse all the files belonging to the profile to insert all the commands to be checked. In this phase,
    we filter global commands and other commands that we know are mandatory. We also have other filters to skip lines or 
    whole SECTIONS (everything indented below something). This should provide a good list with no missing commands, we
    should also update the counters matching each command as we proceed parsing all the files. This should provide a
    confidence level of how much a certain command should be present or not. We store everything in an excel file for
    further manuale processing. As a last step, we parse the 'target' example file for the given profile, and print all
    the missing commands (for example, interfaces configurations, router bgp configurations and so on). These commands
    could be manually changed to be checked on ALL node's configuration (for example 'router bgp\|timers 10 30'), or they
    could be simply removed (like for example all interface specific configs, unless they are used for the same purpose so 
    that the same interfaces is configured in the same way on every node ... ip address config could be removed).
    """
    global commands

    with open(filename, 'r') as file:
        print('Reading file ' + filename + ' ... ')
        last_spaces = 0
        last_line = ""
        parents = []
        num_spaces = []
        spaces = 0
        # reading file line by line
        for line in file:
            if re.search('^\s', line):
                spaces = len(re.search("^(\s+)", line).group(1))
                #print('\n\nEntering line with trailing spaces '+str(spaces)+' last spaces '+str(last_spaces))
                if (spaces > last_spaces):
                    parents.append(last_line.strip())
                    num_spaces.append(spaces)
                    last_spaces = spaces
                    #print('Inside indentation '+last_line+' spaces '+str(spaces)+' last spaces '+str(last_spaces))
                elif (spaces < last_spaces):
                    #print('Exiting indentation with line '+line+' spaces '+str(spaces)+' last spaces '+str(last_spaces)+' num_spaces[-1] '+str(num_spaces[-1]))
                    #print(parents)
                    #print(num_spaces)
                    while (len(num_spaces)>0) and num_spaces[-1]>spaces:
                        #print('Removing from parent stack '+parents[-1])
                        num_spaces.pop()
                        parents.pop()
                        #print(parents)
                    if not len(num_spaces):
                        print('WARNING: indentation error at line '+line+' spaces '+str(spaces)+' last spaces '+str(last_spaces))
                    last_spaces = spaces
            # in this case there are 0 spaces, this is a global config command. Often on Cisco configuration,
            # this line can be a "!" line. This works anyway.
            else:
                while (len(parents)):
                    parents.pop()
                    num_spaces.pop()
                    last_spaces = 0
                    spaces = 0
                    last_line = ''
            last_line = line
            line = get_line(line)
            if line == None:
                continue

            last_line = line
            if len(parents):
                if re.search('( key \d+|enable secret|password)', line):
                    cmd = line_break.join(parents) + line_break + line.strip()
                else:
                    cmd = line_break.join(parents) + line_break + line.strip() + '$'
                if not target and re.search('^(router|interface|crypto pki certificate chain|vrf definition)', cmd):
                    continue
                if not cmd in commands:
                    commands[cmd] = 0
                    #print('"' + '\|'.join(parents) + '\|' + line.strip() + '$",')
                if not target:
                    commands[cmd] += 1
            else:
                if re.search('( key \d+|enable secret \d+|password)', line):
                    cmd = line.strip()
                else:
                    cmd = line.strip() + '$'
                if not target and re.search('^(router|interface|crypto pki certificate chain|vrf definition)', cmd):
                    continue
                if not cmd in commands:
                    commands[cmd] = 0
                    #print('"' + line.strip() + '$",')
                if not target:
                    commands[cmd] += 1

if gen_template:
    """ 
    We start reading the excel file with the full list of devices, we store the lines that belong to the same
    profiles in a list. We can then, for every profile:
    - read the profile command list
    - create the profile sheet (if not existent)
    - parse the configuration file for every device in the list
    - write the output, command present or not, and update its counter
    """
    profiles_list = {}
    dev_targ = openpyxl.load_workbook(xls_ip_devices)
    dev_sheet = dev_targ['Devices']
    for xls_row in range(2, dev_sheet.max_row+1):
        profile = str(dev_sheet['P'+str(xls_row)].value).strip()
        dir = str(dev_sheet['A'+str(xls_row)].value).strip()+'/'
        dev = str(dev_sheet['E'+str(xls_row)].value).strip()
        if not os.path.exists(cfg_root_dir + dir + dev + '.txt'):
            # print('Config not found: "'+cfg_root_dir + dir + dev+'.txt"')
            continue
        if profile == 'None' or len(profile)==0 or not re.search(profiles_filter, profile):
            continue
        if not profile in profiles_list:
            profiles_list[profile] = []
        profiles_list[profile].append(xls_row)

    if os.path.exists(cfg_check_cmd):
        targ = openpyxl.load_workbook(cfg_check_cmd)
    else:
        targ = openpyxl.Workbook()

    for profile in profiles_list:
        if profile in targ:
            sheet = targ[profile]
        else:
            sheet = targ.create_sheet(profile)
        if not profile in cfg_template:
            print('ERROR for profile '+profile+', could not find a template defined')
            continue

        commands = {}
        row_counter = 2
        for i in range(len(profiles_list[profile])):
            xls_row = profiles_list[profile][i]
            dir = str(dev_sheet['A'+str(xls_row)].value).strip() + '/'
            dev = str(dev_sheet['E'+str(xls_row)].value).strip()
            update_template_cmd(cfg_root_dir + dir + dev + '.txt')
        
        # Here we have read all the config files, now it's time to write down the commands in the excel file.
        # We parse one configuration file, and write down the lines that we have already found also on the other
        # nodes. The other lines are printed afterwards.
        update_template_cmd(cfg_template[profile], target = True)
        
        # finished reading the file, here we should print the remained commands, those that should be present only on
        # a few routers. commands[cmd] contains the number of occurrences of the command, on all config files.
        xls_row = 2
        for cmd in commands:
            sheet['A'+str(xls_row)].value = cmd
            sheet['C'+str(xls_row)].value = commands[cmd]
            xls_row += 1
        targ.save(cfg_check_cmd)
        targ.close()

if gen_excel:
    """
    - we read all the devices belonging to all profiles from the main device database, filters
    about profiles and devices are already applied in this phase
    - we open the 'xls_cfg_miss' file and for every profile, we create a new sheet if
    necessary, and we start examining all the devices' config, for each line we check
    if the required command is present or not, with complexity O(Nsquare).
    """
    # this is a dictionary, the keys being the profiles. The values are
    # arrays containing the indexes of the excel database file.
    profiles_list = {}
    dev_targ = openpyxl.load_workbook(xls_ip_devices)
    dev_sheet = dev_targ['Devices']
    for xls_row in range(2, dev_sheet.max_row+1):
        profile = str(dev_sheet['P'+str(xls_row)].value).strip()
        dir = str(dev_sheet['A'+str(xls_row)].value).strip()+'/'
        dev = str(dev_sheet['E'+str(xls_row)].value).strip()
        if not os.path.exists(cfg_root_dir + dir + dev + '.txt'):
            # print('Config not found: "'+cfg_root_dir + dir + dev+'.txt"')
            continue
        if profile == 'None' or len(profile)==0 or not re.search(profiles_filter, profile):
            continue
        if not re.search(dev_filter, dev):
            continue
        if not profile in profiles_list:
            profiles_list[profile] = []
        profiles_list[profile].append(xls_row)
    dev_targ.close()

    if os.path.exists(xls_cfg_miss):
        targ = openpyxl.load_workbook(xls_cfg_miss)
    else:
        targ = openpyxl.Workbook()

    for profile in profiles_list:
        if profile in targ:
            sheet = targ[profile]
        else:
            sheet = targ.create_sheet(profile)

        commands = []
        try:
            cmd_targ = openpyxl.load_workbook(cfg_check_cmd)
            cmd_sheet = cmd_targ[profile]
            for cmd_row in range(2, cmd_sheet.max_row+1):
                cmd_temp = str(cmd_sheet['A'+str(cmd_row)].value).strip()
                if len(cmd_temp) and cmd_temp!='None':
                    commands.append(str(cmd_sheet['A'+str(cmd_row)].value).strip())
        except Exception as err:
            print('Exception while reading commands file '+str(err)+' for profile '+profile)
            continue
        
        for i in range(0, len(commands)):
            cell = commands[i].replace(line_break, '\n')
            sheet[get_column_letter(i+2)+'1'].value = cell

        cmd_counter = [0] * len(commands)
        row_counter = 2
        for i in range(len(profiles_list[profile])):
            xls_row = profiles_list[profile][i]
            dir = str(dev_sheet['A'+str(xls_row)].value).strip() + '/'
            dev = str(dev_sheet['E'+str(xls_row)].value).strip()
            filename = cfg_root_dir + dir + dev + '.txt'
            with open(filename, 'r') as file:
                print('Reading file ' + filename + ' ... ')
                sheet['A'+str(row_counter)].value = os.path.basename(filename).replace('.txt','')
                last_spaces = 0
                last_line = ""
                parents = []
                num_spaces = []
                spaces = 0
                search_line = ''
                # reading file line by line
                for line in file:
                    if re.search('^\s', line):
                        spaces = len(re.search("^(\s+)", line).group(1))
                        #print('Entering line with trailing spaces '+str(spaces)+' last spaces '+str(last_spaces))
                        if (spaces > last_spaces):
                            parents.append(last_line.strip())
                            num_spaces.append(spaces)
                            last_spaces = spaces
                            #print('Inside indentation '+last_line)
                        elif (spaces < last_spaces):
                            #print('Exiting indentation with line '+line+' spaces '+str(spaces)+' last spaces '+str(last_spaces)+' num_spaces[-1] '+str(num_spaces[-1]))
                            #print(parents)
                            #print(num_spaces)
                            while (len(num_spaces)>0) and num_spaces[-1]>spaces:
                                #print('Removing from parent stack '+parents[-1])
                                parents.pop()
                                num_spaces.pop()
                                #print(parents)
                            if not len(num_spaces):
                                print('WARNING: indentation error at line '+line+' spaces '+spaces+' last spaces '+last_spaces)
                            last_spaces = spaces
                    # in this case there are 0 spaces, this is a global config command. Often on Cisco configuration,
                    # this line can be a "!" line. This works anyway.
                    else:
                        while (len(parents)):
                            parents.pop()
                            num_spaces.pop()
                            last_spaces = 0
                            spaces = 0
                            last_line = ''
                    if len(parents):
                        search_line = line_break.join(parents) + line_break + line.strip()
                    else:
                        search_line = line.strip()
                    #print(search_line)
                    #print(parents)
                    for i in range(0, len(commands)):
                        col_string = get_column_letter(i+2)
                        commands[i] = commands[i].replace('***',' ')
                        if re.search(commands[i], search_line) and (commands[i].count(line_break) == search_line.count(line_break)):
                            sheet[col_string + str(row_counter)].value = 'X'
                            cmd_counter[i] += 1
                            break
                    last_line = line.strip()
                # print('Reading '+filename+' ...')
                row_counter += 1

        # print the total number of configs that have that command, it's basicly the number of 'X' in every column
        for i in range(0, len(cmd_counter)):
            cell = get_column_letter(i+2)+str(row_counter)
            total = len(profiles_list[profile])
            sheet[cell].value = str(cmd_counter[i]) + ' / ' + str(total)
            if cmd_counter[i] == total:
                sheet[cell].fill = PatternFill("solid", fgColor="00FF00")
        
        print('Writing '+str(sheet.max_row)+' rows of the excel file for profile '+profile)
        targ.save(xls_cfg_miss)
        targ.close()

def fill_cmd_with_vars (dev, cmd, vars):
    """
    Parameters
    ----------
    dev : str
        the name of the device taken from the excel database
    cmd : str
        The sound the animal makes
    vars : dictionary of dictionaries
        contains the device->var_name->var_value read in the VARS sheet

    Find the occurrences of all the variables written in the "$(var_name)" format and replace them.
    In case there are ip addresses listed, we allow simple operations like addition/subtraction, e.g.:
    $(mgmt_ip) + 1 = 10.10.0.3 + 1 = 10.10.0.4
    This can be useful for static routes with the next hop being based on the interface's ip, that
    has been defined as a variable in the proper tab.
    """
    if not re.search('\$\(.*?\)', cmd):
        return cmd
    if not dev in vars:
        return None
    all_vars = re.findall('\$\((.*?)\)', cmd)
    for var in all_vars:
        if not var in vars[dev]:
            print('ERROR, could not find value for var '+var+' for device '+dev)
            return None
        else:
            cmd = cmd.replace('$('+var+')', vars[dev][var])
    
    # Ip addresses retrieval, replacement and math operations
    if re.search('\d+\.\d+\.\d+\.\d+\s*[+|-]\s*\d+', cmd):
        all_vars = re.findall('((\d+\.\d+\.\d+\.\d+)\s*?([+|-])\s*(\d+))', cmd)
        for var in all_vars:
            try:
                if var[2]=='+':
                    ip = IPv4Address(var[1]) + (int)(var[3])
                else:
                    ip = IPv4Address(var[1]) - (int)(var[3])
                cmd = cmd.replace(var[0], str(ip))
            except ValueError:
                print("Device "+dev+" detected something that looks like an ip address, but it's not: " + var[1])
    return cmd


if fix_cfg:
    """ to fix stuff, we read the output excel file and check the columns, and the presence of the config command.
    We read the cfg_check_cmd file and:
    - if there is a 'add command' in column 'B', we simply add this command in case it's missing
    - if there is a 'change command' in column 'C', it is executed in case the command is PRESENT in the config

    In both cases, specific per-device variables are allowed and can be used with the format $(var_name) and are
    loaded from the 'VARS' sheet.
    
    The above approach is due to the fact that if a command is not present but we used a regexp to make the check, we can't
    replace it with something known, so it must be specified.
    """
    print('Loading commands to be used for the configuration fixes')
    prof_fix_add_cmd = {}
    prof_fix_rem_cmd = {}
    targ = openpyxl.load_workbook(cfg_check_cmd)
    sheet_names = targ.sheetnames
    for profile_sheet in sheet_names:
        if not profile_sheet in prof_fix_add_cmd:
            prof_fix_add_cmd[profile_sheet] = {}
            prof_fix_rem_cmd[profile_sheet] = {}
        if re.search(profiles_filter, profile_sheet):
            sheet = targ[profile_sheet]
            for xls_row in range(2, sheet.max_row+1):
                orig_cmd = str(sheet['A'+str(xls_row)].value).strip().replace(line_break,'\n')
                if sheet['B'+str(xls_row)].value != None and len(str(sheet['B'+str(xls_row)].value).strip()):
                    if sheet['C'+str(xls_row)].value != None and len(str(sheet['C'+str(xls_row)].value).strip()):
                        print('ERROR on line '+str(xls_row)+' profile '+profile_sheet+" columns B and C can't be both full")
                        continue
                    new_cmd = str(sheet['B'+str(xls_row)].value).strip().replace(line_break,'\n')
                    new_cmd = re.sub('\$$', '', new_cmd)
                    prof_fix_add_cmd[profile_sheet][orig_cmd] = new_cmd
                    #print(sheet['B'+str(xls_row)].value)
                if sheet['C'+str(xls_row)].value != None and len(str(sheet['C'+str(xls_row)].value).strip()):
                    new_cmd = str(sheet['C'+str(xls_row)].value).strip().replace(line_break,'\n')
                    new_cmd = re.sub('\$$', '', new_cmd)
                    prof_fix_rem_cmd[profile_sheet][orig_cmd] = new_cmd

    # let's read all the device specific variables and their values
    vars = {}
    sheet = targ['VARS']
    for xls_row in range(2, sheet.max_row+1):
        dev = str(sheet['A'+str(xls_row)].value).strip()
        var_name = str(sheet['B'+str(xls_row)].value).strip()
        var_value = str(sheet['C'+str(xls_row)].value)
        if len(dev):
            if not dev in vars:
                vars[dev] = {}
            vars[dev][var_name] = var_value
    # print(vars)
    targ.close()
    #print(prof_fix_add_cmd)
    #print(prof_fix_rem_cmd)

    # now we should store the rows for all the devices' names, and then we can start parsing the output file with the
    # potentially missing configurations
    print('Loading devices connection details ... ')
    dev_targ = openpyxl.load_workbook(xls_ip_devices)
    dev_sheet = dev_targ['Devices']
    dev_list = {}
    for xls_row in range(2, dev_sheet.max_row+1):
        profile = str(dev_sheet['P'+str(xls_row)].value).strip()
        dev = str(dev_sheet['E'+str(xls_row)].value).strip()
        if profile == 'None' or len(profile) == 0 or not re.search(profiles_filter, profile):
            continue
        if dev in dev_list:
            print('Duplicated device name for '+dev+', we skip it')
            continue
        dev_list[dev] = xls_row

    print('Reading the fixing config file ... ')
    devices_commands = {}
    cmd_targ = openpyxl.load_workbook(xls_cfg_miss)
    sheet_names = cmd_targ.sheetnames
    for profile in sheet_names:
        if not re.search(profiles_filter, profile):
            continue
        cmd_sheet = cmd_targ[profile]
        print('Reading the '+profile+' tab, rows '+str(cmd_sheet.max_row)+' columns '+str(cmd_sheet.max_column)+' ... ')
        for cmd_row in range(2, cmd_sheet.max_row+1):
            if (cmd_row%100 == 0):
                print(' ... read '+str(cmd_row)+' lines')
            dev = str(cmd_sheet['A'+str(cmd_row)].value).strip()
            if not re.search(dev_filter, dev) or dev == 'None':
                continue
            #print('Row '+str(cmd_row)+' for device '+dev)
            if not dev in dev_list:
                print("ERROR couldn't find device "+dev+" in the devices list row "+str(cmd_row)+", skipping it")
                continue
            if not dev in devices_commands:
                devices_commands[dev] = []
            # Now we cycle on the columns. We have basicly two types of commands:
            # 1 - missing commands that need to be added on devices where they are missing
            # 2 - wrong commands that need to be removed/cleaned/changed when they are present
            for cmd_col in range(2, cmd_sheet.max_column+1):
                #print(str(cmd_sheet[get_column_letter(cmd_col)+str(cmd_row)].value))
                cmd_2_check = str(cmd_sheet[get_column_letter(cmd_col)+'1'].value).strip()
                if cmd_sheet[get_column_letter(cmd_col)+str(cmd_row)].value == None:
                    #print('Checking command "'+cmd_2_check+'"')
                    if not cmd_2_check in prof_fix_add_cmd[profile]:
                        continue
                    # We avoid using string.replace because it could lead to unexpected results, we use again
                    # a regular expression, and we replace only strings ending with a dollar. 
                    cmd = re.sub('\$$', '', prof_fix_add_cmd[profile][cmd_2_check])
                    cmd = fill_cmd_with_vars(dev, cmd, vars)
                    if cmd == None:
                        continue
                    #print('Found fix command '+cmd)
                    cmd_list = cmd.split('\n')
                    if len(cmd_list)>1:
                        for k in range(len(cmd_list)-1):
                            cmd_list.append('exit')
                    for k in range(0, len(cmd_list)):
                        devices_commands[dev].append(cmd_list[k])
                else:
                    #print('Checking command "'+cmd_2_check+'"')
                    if not cmd_2_check in prof_fix_rem_cmd[profile]:
                        continue
                    # We avoid using string.replace because it could lead to unexpected results, we use again
                    # a regular expression, and we replace only strings ending with a dollar. 
                    cmd = re.sub('\$$', '', prof_fix_rem_cmd[profile][cmd_2_check])
                    cmd = fill_cmd_with_vars(dev, cmd, vars)
                    if cmd == None:
                        continue
                    # print('Found fix command '+cmd)
                    cmd_list = cmd.split('\n')
                    if len(cmd_list)>1:
                        for k in range(len(cmd_list)-1):
                            cmd_list.append('exit')
                    for k in range(0, len(cmd_list)):
                        devices_commands[dev].append(cmd_list[k])
    
    print('\nList of commands to be executed on '+str(len(devices_commands))+' devices ...\n')
    for dev in devices_commands:
        if len(devices_commands[dev]):
            print('\n\nCommands on '+dev+':')
            print(devices_commands[dev])

    print("Total time:" , datetime.datetime.now()-start)
